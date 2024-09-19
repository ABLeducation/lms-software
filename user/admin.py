from django.contrib import admin
from .models import *
from django.utils.translation import gettext_lazy as _
import openpyxl # type: ignore
from django.http import HttpResponse

admin.site.register(CustomUser)
admin.site.register(Student)
admin.site.register(Teacher)
admin.site.register(School)
# admin.site.register(UserLoginActivity)
admin.site.register(UserActivity1)

class SchoolFilter(admin.SimpleListFilter):
    title = _('school')
    parameter_name = 'school'

    def lookups(self, request, model_admin):
        schools = set()
        for activity in UserLoginActivity.objects.all():
            username = activity.login_username
            if username:
                school = self.get_school_from_username(username)
                if school:
                    schools.add(school)
        return [(school, school) for school in schools]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in Student.objects.filter(school__icontains=self.value())
            ])
        return queryset

    def get_school_from_username(self, username):
        try:
            user_profile = Student.objects.get(user__username=username)
            return user_profile.school
        except Student.DoesNotExist:
            return None
        
class ClassFilter(admin.SimpleListFilter):
    title = _('class')
    parameter_name = 'class'

    def lookups(self, request, model_admin):
        grades = set()
        for student in Student.objects.all():
            grade = student.grade
            if grade:
                grades.add(grade)
        return [(grade, grade) for grade in grades]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in Student.objects.filter(grade__icontains=self.value())
            ])
        return queryset

class SectionFilter(admin.SimpleListFilter):
    title = _('section')
    parameter_name = 'section'

    def lookups(self, request, model_admin):
        sections = set()
        for student in Student.objects.all():
            section = student.section
            if section:
                sections.add(section)
        return [(section, section) for section in sections]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(login_username__in=[
                user.user.username for user in Student.objects.filter(section__icontains=self.value())
            ])
        return queryset

@admin.register(UserLoginActivity)
class UserLoginActivityAdmin(admin.ModelAdmin):
    list_display = ('login_username', 'get_student_name','get_grade','get_section','login_datetime','login_num')
    list_filter = (SchoolFilter, ClassFilter, SectionFilter)
    actions = ['export_as_excel']
    search_fields=('login_username',)

    def get_student_name(self, obj):
        try:
            student_profile = Student.objects.get(user__username=obj.login_username)
            return f"{student_profile.name}"
        except Student.DoesNotExist:
            return None

    get_student_name.short_description = 'Student Name'
    
    def get_grade(self, obj):
        try:
            student_profile = Student.objects.get(user__username=obj.login_username)
            grade=student_profile.grade
            return f"{grade}"
        except Student.DoesNotExist:
            return None
    get_grade.short_description = 'Grade'
        
    def get_section(self, obj):
        try:
            student_profile = Student.objects.get(user__username=obj.login_username)
            section=student_profile.section
            return f"{section}"
        except Student.DoesNotExist:
            return None

    get_section.short_description = 'Section'

    def export_as_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=UserLoginActivities.xlsx'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'User Login Activities'

        # Define the columns
        columns = ['Username', 'Student Name', 'Login Number', 'Login DateTime']

        # Write the header row
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            row = [
                obj.login_username,
                self.get_student_name(obj),
                obj.login_num,
                obj.login_datetime,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    export_as_excel.short_description = "Export Selected as Excel"